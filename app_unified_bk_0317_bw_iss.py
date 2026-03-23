from __future__ import annotations

from flask import Flask, render_template, request, send_file, redirect, url_for, session, jsonify
from io import BytesIO
import datetime as dt
import re
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF
from reportlab.lib import colors
import pandas as pd
import os
import traceback
import tempfile

try:
    import cairosvg
except Exception:
    cairosvg = None

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-only-secret')

TIMING_APP_URL = os.environ.get('TIMING_APP_URL', '').strip()

def get_timing_app_url(req_host:str | None = None) -> str:
    url = (os.environ.get('TIMING_APP_URL') or TIMING_APP_URL or '').strip()
    if not url:
        url = TIMING_APP_URL
        host = (req_host or '').strip()
        if host:
            host = host.split(':',1)[0]
        else:
            host = '127.0.0.1'
        url = f'http://{host}:5001'
    if not re.match(r'^https?://', url):
        url = 'http://' + url.lstrip('/')
    return url.rstrip('/')

# ------------------------- eCPRI 분석 함수 -------------------------
def analyze_ecpri_data(file_path):
    """eCPRI 데이터 분석 함수"""
    results = {}
    print(f"Attempting to analyze file: {file_path}")
    try:
        if not os.path.exists(file_path): 
            return {"error": f"File not found: {file_path}"}
        if not os.path.isfile(file_path): 
            return {"error": f"Path is not a file: {file_path}"}

        print("Loading CSV (skiprows=1, delimiter=';')...")
        try:
            df = pd.read_csv(file_path, delimiter=';', header=0, skiprows=1, low_memory=False)
            print(f"Loaded {len(df)} rows. Columns: {list(df.columns)}")
        except Exception as e:
            return {"error": f"Failed to read CSV (skiprows=1). Check format/delimiter. Error: {e}"}

        print("Cleaning data...")
        required_columns = ['ecpri.dataDir', 'ecpri.message', 'ecpri.rtcId', 'eCpriDelayPtpUs']
        filter_columns = ['ecpri.dataDir', 'ecpri.message', 'ecpri.rtcId']
        delay_col = 'eCpriDelayPtpUs'

        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols: 
            return {"error": f"Missing columns: {', '.join(missing_cols)}."}

        df = df.copy()
        # Convert filter columns to numeric first
        for col in filter_columns:
            if df[col].dtype == 'object': 
                df[col] = df[col].astype(str).str.replace(',', '.', regex=False)
            df[col] = pd.to_numeric(df[col], errors='coerce')
            print(f"Filter Col '{col}' NaNs after coerce: {df[col].isnull().sum()}")

        # Only attempt delay col conversion if it's object type
        if df[delay_col].dtype == 'object':
            df[delay_col] = df[delay_col].astype(str).str.replace(',', '.', regex=False)

        initial_rows = len(df)
        df.dropna(subset=filter_columns, inplace=True)
        print(f"Dropped {initial_rows - len(df)} rows due to NaNs in filter columns.")

        if df.empty: 
            return {"error": "No valid data after cleaning filter columns."}

        print("Identifying top 4 rtcIds for UL...")
        
        # UL 데이터에서 상위 4개 rtcId 찾기
        ul_control_condition = (df['ecpri.dataDir'] == 0) & (df['ecpri.message'] == 2)
        ul_control_data = df.loc[ul_control_condition].copy()

        top_4_rtcId_ul = []
        if not ul_control_data.empty and 'ecpri.rtcId' in ul_control_data.columns:
            ul_control_rtc_clean = ul_control_data['ecpri.rtcId'].dropna()
            if not ul_control_rtc_clean.empty:
                # 정수형으로 변환 시도
                try:
                    ul_control_rtc_clean = ul_control_rtc_clean.astype(int)
                    top_4_rtcId_ul = ul_control_rtc_clean.value_counts().nlargest(4).index.tolist()
                except (ValueError, TypeError):
                    # 정수 변환 실패 시 그대로 사용
                    top_4_rtcId_ul = ul_control_rtc_clean.value_counts().nlargest(4).index.tolist()
                print(f"Found {len(top_4_rtcId_ul)} top RTC IDs from Control Plane UL")
            else: 
                print("Warning: No valid rtcId values in Control Plane UL data.")
        else: 
            print("Warning: No Control Plane UL data (dataDir=0, message=2).")

        print(f"Top 4 rtcIds for UL filtering: {top_4_rtcId_ul}")

        print("Defining filters...")
        
        # 필터 정의
        filters = {
            "User plane DL": (df['ecpri.dataDir'] == 1) & (df['ecpri.message'] == 0),
            "Control Plane DL": (df['ecpri.dataDir'] == 1) & (df['ecpri.message'] == 2),
            "Control Plane UL": (df['ecpri.dataDir'] == 0) & (df['ecpri.message'] == 2),
            "User plane UL": (df['ecpri.dataDir'] == 0) & (df['ecpri.message'] == 0)
        }

        # RTC ID 필터링이 가능한 경우에만 적용
        if top_4_rtcId_ul:
            print("Applying RTC ID filtering for UL data...")
            filters["Control Plane UL"] = filters["Control Plane UL"] & (df['ecpri.rtcId'].isin(top_4_rtcId_ul))
            filters["User plane UL"] = filters["User plane UL"] & (df['ecpri.rtcId'].isin(top_4_rtcId_ul))
        else:
            print("No RTC ID filtering applied for UL data")

        print("Applying filters and calculating min/max...")
        
        # 각 카테고리별 통계 계산
        for name, condition in filters.items():
            try:
                filtered_df = df.loc[condition]
                count = len(filtered_df)

                min_val, max_val = "N/A", "N/A"
                valid_delay_count = 0

                if not filtered_df.empty:
                    delay_values_numeric = pd.to_numeric(filtered_df[delay_col], errors='coerce')
                    valid_delay_values = delay_values_numeric.dropna()
                    valid_delay_count = len(valid_delay_values)

                    if valid_delay_count > 0:
                        min_val = float(valid_delay_values.min())
                        max_val = float(valid_delay_values.max())
                        print(f"Category '{name}': Found {count} rows. Min/Max on {valid_delay_count} values. Min: {min_val:.2f}, Max: {max_val:.2f}")
                    else:
                        print(f"Category '{name}': Found {count} rows, but no valid numeric '{delay_col}'.")
                else:
                    print(f"Category '{name}': Found 0 rows.")

                results[name] = {"min": min_val, "max": max_val, "count": count}

            except Exception as filter_err:
                print(f"Error processing category '{name}': {filter_err}")
                print(traceback.format_exc())
                results[name] = {"min": "Error", "max": "Error", "count": 0}

        print("Analysis function finished.")
        return results

    except Exception as e:
        print("--- Detailed Error Traceback ---")
        print(traceback.format_exc())
        return {"error": f"Unexpected analysis error: {str(e)}"}

def analyze_iq_data(df):
    """IQ 데이터 분석"""
    results = {}
    
    try:
        # IQ 데이터 관련 컬럼 확인
        iq_columns = [col for col in df.columns if 'iq' in col.lower() or 'sample' in col.lower()]
        
        if not iq_columns:
            results['info'] = "No IQ-specific columns found in the dataset"
            return results
        
        # IQ 데이터 기본 통계
        for col in iq_columns[:5]:  # 처음 5개 컬럼만 분석
            if df[col].dtype in ['float64', 'int64']:
                numeric_data = pd.to_numeric(df[col], errors='coerce').dropna()
                if not numeric_data.empty:
                    results[col] = {
                        'min': float(numeric_data.min()),
                        'max': float(numeric_data.max()),
                        'mean': float(numeric_data.mean()),
                        'std': float(numeric_data.std()),
                        'count': len(numeric_data)
                    }
        
        # IQ 데이터 분포 분석
        results['iq_analysis'] = {
            'total_iq_columns': len(iq_columns),
            'analyzed_columns': min(5, len(iq_columns)),
            'total_samples': len(df)
        }
        
    except Exception as e:
        print(f"IQ data analysis error: {e}")
        
    return results

def build_basic_csv_stats(file_path):
    """CSV 기본 통계 정보 생성"""
    df = pd.read_csv(file_path, delimiter=';', header=0, skiprows=1, low_memory=False)
    return {
        'total_rows': len(df),
        'columns': list(df.columns),
        'data_types': {col: str(dtype) for col, dtype in df.dtypes.items()}
    }, df

# ------------------------- Excel 출력 기능 -------------------------
def save_ecpri_results_to_excel(results, filename=None):
    """eCPRI 분석 결과를 Excel 파일로 저장"""
    if filename is None:
        filename = f"ecpri_analysis_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 결과 데이터 준비 - 요청한 형식으로
    excel_data = []
    categories = ["User plane DL", "Control Plane DL", "Control Plane UL", "User plane UL"]
    
    for category in categories:
        if category in results:
            data = results[category]
            # Min Delay 행 추가
            min_value = data["min"] 
            if isinstance(data["min"], (int, float)):
                min_value = f"{min_value:.2f}"
            else:
                min_value = "N/A"
            excel_data.append({
                "Category": category,
                "Metric": "Min Delay (µs)",
                "Value": min_value
            })
            # Max Delay 행 추가
            max_value = data["max"] 
            if isinstance(data["max"], (int, float)):
                max_value = f"{max_value:.2f}"
            else:
                max_value = "N/A"
            excel_data.append({
                "Category": category,
                "Metric": "Max Delay (µs)", 
                "Value": max_value
            })
    
    # DataFrame 생성
    df = pd.DataFrame(excel_data)
    
    # Excel 파일로 저장
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='eCPRI Analysis', index=False)
        
        # 워크시트 스타일링
        workbook = writer.book
        worksheet = writer.sheets['eCPRI Analysis']
        
        # 컬럼 너비 조정
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 12
        
        # 헤더 스타일링
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    output.seek(0)
    return output, filename

# ------------------------- 기존 DL/UL 함수들 -------------------------
def nnum(v):
    """숫자 변환 (실패 시 None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\u202f", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def fmt_ns(v):
    """타임라인용 ns 포맷."""
    if v is None:
        return "N/A"
    try:
        f = float(v)
    except Exception:
        return str(v)
    return f"{f:.1f}"

def normalize_key(k: str) -> str:
    """엑셀 키에서 '(ns)' 같은 것 제거."""
    return re.sub(r"\s*\(ns\)\s*:?$", "", str(k)).strip()

def fmt_cell(v):
    """PDF용 간단 포맷."""
    if v is None:
        return ""
    try:
        f = float(v)
    except Exception:
        return str(v)
    if f.is_integer():
        return str(int(f))
    return f"{f:.3f}".rstrip("0").rstrip(".")

@app.template_filter("fmt")
def jfmt(v):
    """Jinja 필터: {{ value|fmt }}."""
    try:
        if v is None:
            return ""
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return f"{f:.6f}".rstrip("0").rstrip(".")
    except Exception:
        return str(v)

# ------------------------- 파라미터 정의 -------------------------
PARAM_DEFAULTS_DL = {
    "T1a_max_cp_dl": -5000.0,
    "T1a_min_cp_dl": 5000.0,
    "T1a_max_up": -3000.0,
    "T1a_min_up": 3000.0,
    "T2a_max_cp_dl": -4500.0,
    "T2a_min_cp_dl": 4500.0,
    "T2a_max_up": -2800.0,
    "T2a_min_up": 2800.0,
    "real_T1a_max_cp_dl": -4800.0,
    "real_T1a_min_cp_dl": 4700.0,
    "real_T1a_max_up": -3100.0,
    "real_T1a_min_up": 3100.0,
    "real_T2a_max_cp_dl": -4600.0,
    "real_T2a_min_cp_dl": 4550.0,
    "real_T2a_max_up": -2700.0,
    "real_T2a_min_up": 2950.0,
    "color_odu_cp": "#60a5fa",
    "color_odu_up": "#65a30d",
    "color_oru_cp": "#fb923c",
    "color_oru_up": "#7c3aed",
}

PARAM_DEFAULTS_UL = {
    "T1a_max_cp_ul": -601.0,
    "T1a_min_cp_ul": -380.0,
    "Ta4_max_ul": 392.0,
    "Ta4_min_ul": 220.0,
    "T2a_max_cp_ul": -451.0,
    "T2a_min_cp_ul": -220.0,
    "Ta3_max_ul": 232.0,
    "Ta3_min_ul": 70.0,
    "real_T1a_max_cp_ul": -488.0,
    "real_T1a_min_cp_ul": -462.0,
    "real_Ta4_max_ul": 328.0,
    "real_Ta4_min_ul": 276.0,
    "real_T2a_max_cp_ul": -241.0,
    "real_T2a_min_cp_ul": -215.0,
    "real_Ta3_max_ul": 186.0,
    "real_Ta3_min_ul": 126.0,
    "color_odu_cp": "#60a5fa",
    "color_odu_up": "#65a30d",
    "color_oru_cp": "#fb923c",
    "color_oru_up": "#7c3aed",
}

# ------------------------- 윈도우 정의 및 평가 함수 -------------------------
WINDOWS_DL = [
    ("O-DU DL Tx C-Plane Window", "T1a_max_cp_dl", "real_T1a_max_cp_dl", "T1a_min_cp_dl", "real_T1a_min_cp_dl", "O-DU", "C-Plane", "T1a"),
    ("O-DU DL Tx U-Plane Window", "T1a_max_up", "real_T1a_max_up", "T1a_min_up", "real_T1a_min_up", "O-DU", "U-Plane", "T1a"),
    ("O-RU DL Rx C-Plane Window", "T2a_max_cp_dl", "real_T2a_max_cp_dl", "T2a_min_cp_dl", "real_T2a_min_cp_dl", "O-RU", "C-Plane", "T2a"),
    ("O-RU DL Rx U-Plane Window", "T2a_max_up", "real_T2a_max_up", "T2a_min_up", "real_T2a_min_up", "O-RU", "U-Plane", "T2a"),
]

WINDOWS_UL = [
    ("O-DU UL Tx C-Plane Window", "T1a_max_cp_ul", "real_T1a_max_cp_ul", "T1a_min_cp_ul", "real_T1a_min_cp_ul", "O-DU", "C-Plane", False, "T1a"),
    ("O-DU UL Rx U-Plane Window", "Ta4_max_ul", "real_Ta4_max_ul", "Ta4_min_ul", "real_Ta4_min_ul", "O-DU", "U-Plane", True, "Ta4"),
    ("O-RU UL Rx C-Plane Window", "T2a_max_cp_ul", "real_T2a_max_cp_ul", "T2a_min_cp_ul", "real_T2a_min_cp_ul", "O-RU", "C-Plane", False, "T2a"),
    ("O-RU UL Tx U-Plane Window", "Ta3_max_ul", "real_Ta3_max_ul", "Ta3_min_ul", "real_Ta3_min_ul", "O-RU", "U-Plane", True, "Ta3"),
]

# ------------------------- DL/UL 판단 함수 (질문 정의 반영) -------------------------

def dl_start_status(meas_start, exp_max_neg, exp_min_neg):
    """
    DL Start 판단 (음수 지연 기준)

    Exp max(-) >  Meas Start(-): Early
    Exp min(-) <  Meas Start(-): Late
    Other : PASS
    """
    if meas_start is None or exp_max_neg is None or exp_min_neg is None:
        return "N/A", "#9ca3af"

    # Early: Exp max(-) > Meas Start(-)
    if exp_max_neg > meas_start:
        return "EARLY", "#b45309"

    # Late: Exp min(-) < Meas Start(-)
    if exp_min_neg < meas_start:
        return "LATE", "#dc2626"

    # 나머지 PASS
    return "PASS", "#16a34a"

def dl_end_status(meas_end, exp_max_neg, exp_min_neg):
    if meas_end is None or exp_max_neg is None or exp_min_neg is None:
        return "N/A", "#9ca3af"

    # 윈도우 내부: PASS
    window_min = min(exp_max_neg, exp_min_neg)
    window_max = max(exp_max_neg, exp_min_neg)
    if window_min <= meas_end <= window_max:
        return "PASS", "#16a34a"

    # 윈도우보다 더 음수(왼쪽)면 EARLY
    if meas_end < window_min:
        return "EARLY", "#b45309"

    # 윈도우보다 덜 음수(오른쪽)면 LATE
    return "LATE", "#dc2626"

def evaluate_dl(params: dict):
    """
    DL 윈도우 평가 (질문에서 정의한 음수 조건 사용)
    WINDOWS_DL:
      (label, exp_max_key, meas_start_key, exp_min_key, meas_end_key, equip, plane, timer_name)
    """
    evals = []
    counters = {"pass": 0, "fail": 0, "ontime": 0, "early": 0, "late": 0}
    log_lines = []

    for window_info in WINDOWS_DL:
        if len(window_info) >= 8:
            label, esK, msK, eeK, meK, equip, plane, timer_name = window_info
        else:
            label, esK, msK, eeK, meK, equip, plane = window_info[:7]
            timer_name = ""

        # 질문 정의에 맞게: esK = Exp max(-), eeK = Exp min(-)
        exp_max_neg = nnum(params.get(esK))
        exp_min_neg = nnum(params.get(eeK))
        meas_start = nnum(params.get(msK))
        meas_end = nnum(params.get(meK))

        s_txt, _ = dl_start_status(meas_start, exp_max_neg, exp_min_neg)
        e_txt, _ = dl_end_status(meas_end, exp_max_neg, exp_min_neg)

        if s_txt == "PASS" and e_txt == "PASS":
            overall = "PASS"
        elif s_txt == "N/A" or e_txt == "N/A":
            overall = "N/A"
        else:
            overall = "FAIL"

        evals.append({
            "label": label,
            "equip": equip,
            "plane": plane,
            "es": exp_max_neg,
            "ms": meas_start,
            "ee": exp_min_neg,
            "me": meas_end,
            "start_status": s_txt,
            "end_status": e_txt,
            "overall": overall,
            "timer_name": timer_name,
        })

        if overall == "PASS":
            counters["pass"] += 1
        elif overall == "FAIL":
            counters["fail"] += 1

        if s_txt == "PASS" and e_txt == "PASS":
            counters["ontime"] += 1
        if s_txt == "EARLY":
            counters["early"] += 1
        if e_txt == "LATE":
            counters["late"] += 1

        log_lines.append(
            f"{label} Start: measured={meas_start} , "
            f"Exp_max_neg={exp_max_neg}, Exp_min_neg={exp_min_neg} ⇒ {s_txt}"
        )
        log_lines.append(
            f"{label} End  : measured={meas_end} , "
            f"Exp_max_neg={exp_max_neg}, Exp_min_neg={exp_min_neg} ⇒ {e_txt}"
        )

    # O-RU U-Plane / C-Plane 선택
    rx_u = next((ev for ev in evals if ev["equip"] == "O-RU" and ev["plane"] == "U-Plane"), None)
    rx_c = next((ev for ev in evals if ev["equip"] == "O-RU" and ev["plane"] == "C-Plane"), None)

    def decide_single_signal(ev):
        out = {"on_time": "", "early": "", "late": ""}
        if ev is None:
            return out
        if ev["start_status"] == "EARLY":
            out["early"] = "FAIL"
            return out
        if ev["end_status"] == "LATE":
            out["late"] = "FAIL"
            return out
        if ev["start_status"] == "PASS" and ev["end_status"] == "PASS":
            out["on_time"] = "PASS"
            return out
        return out

    rx_u_sig = decide_single_signal(rx_u)
    rx_c_sig = decide_single_signal(rx_c)

    rx_checks = {
        "RX_On_Time":   rx_u_sig["on_time"],
        "RX_Early":     rx_u_sig["early"],
        "RX_Late":      rx_u_sig["late"],
        "RX_On_Time_C": rx_c_sig["on_time"],
        "RX_Early_C":   rx_c_sig["early"],
        "RX_Late_C":    rx_c_sig["late"],
    }

    return evals, counters, "\n".join(log_lines), rx_checks

# -------- UL 음수/양수 판단 --------
def ul_status_neg(meas, exp_max_neg, exp_min_neg, is_start: bool):
    """
    UL 판단 (음수 지연 기준)

    Start:
      Exp max(-) >  Meas Start(-) : Early
      Exp min(-) <  Meas Start(-) : Late
      Other : PASS

    End:
      Exp min(-) <  Meas End (-)  : Late
      Exp max(-) > Meas End(-)    : Early
      Other : PASS
    """
    if meas is None or exp_max_neg is None or exp_min_neg is None:
        return "N/A", "#9ca3af"

    if is_start:
        # Early: Exp max(-) > Meas Start(-)
        if exp_max_neg > meas:
            return "EARLY", "#b45309"
        # Late: Exp min(-) < Meas Start(-)
        if exp_min_neg < meas:
            return "LATE", "#dc2626"
        return "PASS", "#16a34a"
    else:
        # End: Late / Early 순서
        # Late: Exp min(-) < Meas End(-)
        if exp_min_neg < meas:
            return "LATE", "#dc2626"
        # Early: Exp max(-) > Meas End(-)
        if exp_max_neg > meas:
            return "EARLY", "#b45309"
        return "PASS", "#16a34a"


def ul_status_pos(meas, exp_max_pos, exp_min_pos, is_start: bool):
    """
    UL 판단 (양수 지연 기준)

    Start(+):
      Exp min(+) > Meas Start (+) : Early
      Exp max(+) < Meas Start(+)  : Late
      Other: PASS

    End(+):
      Exp max(+) <  Meas End (+)  : Late
      Exp min(+) > Meas End(+)    : Early
      Other: PASS
    """
    if meas is None or exp_max_pos is None or exp_min_pos is None:
        return "N/A", "#9ca3af"

    if is_start:
        # Early: Exp min(+) > Meas Start(+)
        if exp_min_pos > meas:
            return "EARLY", "#b45309"
        # Late: Exp max(+) < Meas Start(+)
        if exp_max_pos < meas:
            return "LATE", "#dc2626"
        return "PASS", "#16a34a"
    else:
        # Late: Exp max(+) < Meas End(+)
        if exp_max_pos < meas:
            return "LATE", "#dc2626"
        # Early: Exp min(+) > Meas End(+)
        if exp_min_pos > meas:
            return "EARLY", "#b45309"
        return "PASS", "#16a34a"


def evaluate_ul(params: dict):
    """
    UL 윈도우 평가 (질문 정의의 음수/양수 조건 사용)
    WINDOWS_UL:
      (label, exp_max_key, meas_start_key, exp_min_key, meas_end_key, equip, plane, is_positive, timer_name)
    """
    evals = []
    counters = {"pass": 0, "fail": 0, "ontime": 0, "early": 0, "late": 0}
    log_lines = []

    for window_info in WINDOWS_UL:
        if len(window_info) >= 9:
            label, esK, msK, eeK, meK, equip, plane, is_positive, timer_name = window_info
        else:
            label, esK, msK, eeK, meK, equip, plane = window_info[:7]
            is_positive = False
            timer_name = ""

        exp_max = nnum(params.get(esK))
        exp_min = nnum(params.get(eeK))

        ms_raw = nnum(params.get(msK))
        me_raw = nnum(params.get(meK))

        # is_positive=True 인 윈도우(Ta3/Ta4)는 기존 로직처럼 ms/me 스왑해서 양수 방향으로 취급
        if is_positive:
            meas_start = me_raw
            meas_end = ms_raw
        else:
            meas_start = ms_raw
            meas_end = me_raw

        if is_positive:
            s_txt, s_col = ul_status_pos(meas_start, exp_max, exp_min, is_start=True)
            e_txt, e_col = ul_status_pos(meas_end,   exp_max, exp_min, is_start=False)
        else:
            s_txt, s_col = ul_status_neg(meas_start, exp_max, exp_min, is_start=True)
            e_txt, e_col = ul_status_neg(meas_end,   exp_max, exp_min, is_start=False)

        if s_txt == "PASS" and e_txt == "PASS":
            overall = "PASS"
        elif s_txt == "N/A" or e_txt == "N/A":
            overall = "N/A"
        else:
            overall = "FAIL"

        evals.append({
            "label": label,
            "equip": equip,
            "plane": plane,
            "es": exp_max,
            "ms": meas_start,
            "ee": exp_min,
            "me": meas_end,
            "start_status": s_txt,
            "start_color": s_col,
            "end_status": e_txt,
            "end_color": e_col,
            "overall": overall,
            "is_positive": is_positive,
            "timer_name": timer_name,
        })

        if overall == "PASS":
            counters["pass"] += 1
        elif overall == "FAIL":
            counters["fail"] += 1

        if s_txt == "PASS" and e_txt == "PASS":
            counters["ontime"] += 1
        if s_txt == "EARLY" or e_txt == "EARLY":
            counters["early"] += 1
        if s_txt == "LATE" or e_txt == "LATE":
            counters["late"] += 1

        log_lines.append(
            f"{label} Start: meas={meas_start}, "
            f"Exp_max={exp_max}, Exp_min={exp_min} ⇒ {s_txt}"
        )
        log_lines.append(
            f"{label} End  : meas={meas_end}, "
            f"Exp_max={exp_max}, Exp_min={exp_min} ⇒ {e_txt}"
        )

    rx_u = next((ev for ev in evals if ev["equip"] == "O-RU" and ev["plane"] == "U-Plane"), None)
    rx_c = next((ev for ev in evals if ev["equip"] == "O-RU" and ev["plane"] == "C-Plane"), None)

    def decide_single_signal(ev):
        out = {"on_time": "", "early": "", "late": ""}
        if ev is None:
            return out
        if ev["start_status"] == "LATE" or ev["end_status"] == "LATE":
            out["late"] = "FAIL"
            return out
        if ev["start_status"] == "EARLY" or ev["end_status"] == "EARLY":
            out["early"] = "FAIL"
            return out
        if ev["start_status"] == "PASS" and ev["end_status"] == "PASS":
            out["on_time"] = "PASS"
            return out
        return out

    rx_u_sig = decide_single_signal(rx_u)
    rx_c_sig = decide_single_signal(rx_c)

    rx_checks = {
        "RX_On_Time":   rx_u_sig["on_time"],
        "RX_Early":     rx_u_sig["early"],
        "RX_Late":      rx_u_sig["late"],
        "RX_On_Time_C": rx_c_sig["on_time"],
        "RX_Early_C":   rx_c_sig["early"],
        "RX_Late_C":    rx_c_sig["late"],
    }

    return evals, counters, "\n".join(log_lines), rx_checks
# =============================================================================
# 기존 SVG 함수 호환 별칭 (인자 개수 맞춤!)
# =============================================================================
def start_status(meas, exp_start):  # 2개 인자 ← 3개로 변환
    return dl_start_status(meas, exp_start, exp_start)  # exp_max_neg=exp_min_neg=exp_start

def end_status(meas, exp_end):      # 2개 인자 ← 3개로 변환  
    return dl_end_status(meas, exp_end, exp_end)      # exp_max_neg=exp_min_neg=exp_end

def window_status(meas, exp_max, exp_min):  # 3개 인자 그대로
    return ul_status_neg(meas, exp_max, exp_min, True)
# =============================================================================
# 기존 SVG 함수들(generate_svg_dl/ul)이 호출하는 함수들의 호환 별칭 (500 에러 해결)
# =============================================================================
start_status = dl_start_status
end_status = dl_end_status
window_status = lambda meas, exp_max, exp_min: ul_status_neg(meas, exp_max, exp_min, True)
# ------------------------- SVG 생성 함수 -------------------------
def generate_svg_dl(params: dict) -> str:
    """DL Delay Window SVG 생성."""
    W = 1200
    LEFT_PAD = 180
    PAD_TOP = 40
    ROW_H = 40
    
    # 축 범위 계산
    all_values = []
    for key in ['T1a_max_cp_dl', 'T1a_min_cp_dl', 'T1a_max_up', 'T1a_min_up', 
                'T2a_max_cp_dl', 'T2a_min_cp_dl', 'T2a_max_up', 'T2a_min_up',
                'real_T1a_max_cp_dl', 'real_T1a_min_cp_dl', 'real_T1a_max_up', 'real_T1a_min_up',
                'real_T2a_max_cp_dl', 'real_T2a_min_cp_dl', 'real_T2a_max_up', 'real_T2a_min_up']:
        val = nnum(params.get(key))
        if val is not None:
            all_values.append(val)
    
    if not all_values:
        minX, maxX = -6000, 6000
    else:
        minX, maxX = min(all_values), max(all_values)
        span = maxX - minX
        if span == 0:
            span = 1000
        pad = span * 0.10
        minX -= pad
        maxX += pad
    
    def scale(x):
        return LEFT_PAD + (x - minX) / (maxX - minX) * (W - LEFT_PAD - 80)
    
    pairs = [
        ("O-DU", "Tx Win C-P",
         "T1a_max_cp_dl", "T1a_min_cp_dl",
         "real_T1a_max_cp_dl", "real_T1a_min_cp_dl", "T1a"),
        ("O-DU", "Tx Win U-P",
         "T1a_max_up", "T1a_min_up",
         "real_T1a_max_up", "real_T1a_min_up", "T1a"),
        ("O-RU", "Rx Win C-P",
         "T2a_max_cp_dl", "T2a_min_cp_dl",
         "real_T2a_max_cp_dl", "real_T2a_min_cp_dl", "T2a"),
        ("O-RU", "Rx Win U-P",
         "T2a_max_up", "T2a_min_up",
         "real_T2a_max_up", "real_T2a_min_up", "T2a"),
    ]
    
    H = PAD_TOP + len(pairs) * (ROW_H * 2 + 18) + 40
    parts = []
    parts.append(f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" class="w-full border rounded-xl bg-white">')
    
    # 배경 그리드 및 축
    parts.append(f'<g transform="translate(0,{PAD_TOP})">')
    parts.append(f'<line x1="{LEFT_PAD}" x2="{W-80}" y1="0" y2="0" stroke="#cbd5e1" stroke-width="1"/>')
    
    # 그리드 라인 및 레이블
    for i in range(5):
        t = i / 4.0
        xval = minX + t * (maxX - minX)
        x = scale(xval)
        parts.append(f'<line x1="{x:.2f}" x2="{x:.2f}" y1="0" y2="{H-70}" stroke="#e2e8f0" stroke-width="1"/>')
        parts.append(f'<text x="{x:.2f}" y="-8" text-anchor="middle" font-size="11" fill="#475569">{fmt_ns(xval)}</text>')
    
    # 중앙선 (t=0)
    x0 = scale(0.0)
    parts.append(f'<line x1="{x0:.2f}" x2="{x0:.2f}" y1="0" y2="{H-70}" stroke="#94a3b8" stroke-dasharray="4 4" stroke-width="1"/>')
    parts.append(f'<text x="{x0:.2f}" y="14" font-size="11" fill="#334155">t=0</text>')
    parts.append('</g>')
    
    # 데이터 영역
    parts.append('<g>')
    TEXT_DARK = "#0f172a"
    
    for idx, pair_info in enumerate(pairs):
        equip, lane, k_es, k_ee, k_ms, k_me, timer_name = pair_info
        
        baseY = PAD_TOP + 18 + idx * (ROW_H * 2 + 18)
        
        # 장비 레이블 (C-P일 때만 표시)
        if "C-P" in lane:
            parts.append(f'<text x="{LEFT_PAD-40}" y="{baseY-6}" text-anchor="end" font-size="15" font-weight="600" fill="{TEXT_DARK}">{equip}</text>')
        
        # 값 가져오기
        es = nnum(params.get(k_es))
        ee = nnum(params.get(k_ee))
        ms = nnum(params.get(k_ms))
        me = nnum(params.get(k_me))
        
        # 색상 팔레트
        EXP_FILL, EXP_EDGE, MEA_FILL, MEA_EDGE = palette_for_row(equip, lane, params)
        
        # 평면 타입
        plane_type = "CP" if "C-P" in lane else "UP"
        
        # -------- Expected row --------
        y_exp = baseY + 10
        if es is None or ee is None:
            parts.append(f'<text x="{LEFT_PAD-30}" y="{y_exp+16}" font-size="11" fill="#ef4444">N/A</text>')
        else:
            xs = scale(es)
            xe = scale(ee)
            rectX = min(xs, xe)
            rectW = abs(xe - xs)
            
            # 예상 영역
            parts.append(f'<rect x="{rectX:.2f}" y="{y_exp}" width="{rectW:.2f}" height="24" fill="{EXP_FILL}" stroke="{EXP_EDGE}" stroke-width="1.5"/>')
            
            # 시작 마커
            parts.append(f'<line x1="{xs:.2f}" x2="{xs:.2f}" y1="{y_exp-8}" y2="{y_exp+24}" stroke="{EXP_EDGE}" stroke-width="2"/>')
            parts.append(f'<text x="{xs:.2f}" y="{y_exp-12}" text-anchor="middle" font-size="11" fill="#1e3a8a">Exp {timer_name} {plane_type} max {fmt_ns(es)}</text>')
            
            # 끝 마커
            parts.append(f'<line x1="{xe:.2f}" x2="{xe:.2f}" y1="{y_exp-8}" y2="{y_exp+24}" stroke="{EXP_EDGE}" stroke-width="2"/>')
            parts.append(f'<text x="{xe:.2f}" y="{y_exp-12}" text-anchor="middle" font-size="11" fill="#1e3a8a">Exp {timer_name} {plane_type} min {fmt_ns(ee)}</text>')
            
            # 레이블
            parts.append(f'<text x="{LEFT_PAD-70}" y="{y_exp+18}" font-size="11" fill="{TEXT_DARK}">{lane} Exp</text>')
        
        # -------- Measured row --------
        y_mea = y_exp + ROW_H
        if ms is None or me is None:
            parts.append(f'<text x="{LEFT_PAD-30}" y="{y_mea+16}" font-size="11" fill="#ef4444">N/A</text>')
        else:
            xs_m = scale(ms)
            xe_m = scale(me)
            rectX_m = min(xs_m, xe_m)
            rectW_m = abs(xe_m - xs_m)
            
            # 측정 영역
            parts.append(f'<rect x="{rectX_m:.2f}" y="{y_mea}" width="{rectW_m:.2f}" height="24" fill="{MEA_FILL}" stroke="{MEA_EDGE}" stroke-width="1.5"/>')
            
            # 상태 평가
            s_txt, s_col = dl_start_status(ms, es, ee)
            e_txt, e_col = dl_end_status(me, ee, es)
            
            # 시작 마커
            parts.append(f'<line x1="{xs_m:.2f}" x2="{xs_m:.2f}" y1="{y_mea-8}" y2="{y_mea+24}" stroke="{s_col}" stroke-width="2"/>')
            start_label_x = xs_m - 10
            start_label_y = y_mea + 10
            parts.append(f'<text x="{start_label_x:.2f}" y="{start_label_y:.2f}" text-anchor="end" font-size="11" fill="{s_col}">Start {timer_name} {plane_type} {fmt_ns(ms)} {s_txt}</text>')
            
            # 끝 마커
            parts.append(f'<line x1="{xe_m:.2f}" x2="{xe_m:.2f}" y1="{y_mea-8}" y2="{y_mea+24}" stroke="{e_col}" stroke-width="2"/>')
            end_label_x = xe_m + 10
            end_label_y = y_mea + 10
            parts.append(f'<text x="{end_label_x:.2f}" y="{end_label_y:.2f}" font-size="11" fill="{e_col}">End {timer_name} {plane_type} {fmt_ns(me)} {e_txt}</text>')
            
            # 레이블
            parts.append(f'<text x="{LEFT_PAD-70}" y="{y_mea+18}" font-size="11" fill="{TEXT_DARK}">{lane} Meas</text>')
    
    parts.append('</g>')
    parts.append('</svg>')
    return "".join(parts)

# ------------------------- Profile Extractor 함수 -------------------------
def extract_delay_profile_data(log_file):
    """
    <rpc-reply> 블록에서 Delay Profile 데이터를 추출하여 리스트로 반환합니다.
    """
    results = []
    
    if not os.path.exists(log_file):
        return {"error": f"File not found - {log_file}"}

    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # [Step 1] <rpc-reply> 블록만 분리 (re.DOTALL 사용)
        reply_pattern = re.compile(r'(<[\w\-\:]*rpc-reply.*?>.*?</[\w\-\:]*rpc-reply>)', re.DOTALL)
        replies = reply_pattern.findall(content)

        if not replies:
            return {"error": "No <rpc-reply> blocks found."}

        found_profiles = []

        # [Step 2] 각 reply 블록 내부에서 bandwidth-scs-delay-state 검색
        for reply_block in replies:
            if "bandwidth-scs-delay-state" not in reply_block:
                continue

            state_pattern = re.compile(r'(<[\w\-\:]*bandwidth-scs-delay-state.*?>.*?</[\w\-\:]*bandwidth-scs-delay-state>)', re.DOTALL)
            state_blocks = state_pattern.findall(reply_block)
            found_profiles.extend(state_blocks)

        if not found_profiles:
            return {"error": "No valid Delay Profile data found inside RPC-Replies."}

        # [Step 3] 값 추출
        for i, block in enumerate(found_profiles):
            try:
                def get_val(tag):
                    # 태그 사이의 값만 엄격하게 추출
                    p = re.compile(r'<[\w\-\:]*'+tag+r'.*?>\s*([^<]+)\s*<\/', re.DOTALL)
                    m = p.search(block)
                    return m.group(1).strip() if m else "N/A"

                bw = get_val("bandwidth")
                scs = get_val("subcarrier-spacing")

                # 포맷팅
                if bw.isdigit() and int(bw) >= 1000000:
                    bw_disp = f"{int(bw)//1000000} MHz"
                else:
                    bw_disp = f"{bw} Hz"
                
                scs_disp = f"{scs} Hz"

                keys = [
                    't2a-min-up', 't2a-max-up',
                    't2a-min-cp-dl', 't2a-max-cp-dl',
                    'tcp-adv-dl',
                    'ta3-min', 'ta3-max',
                    't2a-min-cp-ul', 't2a-max-cp-ul'
                ]
                
                row_data = {
                    "id": i + 1,
                    "bandwidth": bw_disp,
                    "scs": scs_disp,
                    "raw_block": block # 필요시 원본 확인용
                }
                
                for k in keys:
                    row_data[k] = get_val(k)
                
                results.append(row_data)

            except Exception as e:
                print(f"Error parsing block {i}: {e}")
                continue
                
        return {"data": results, "count": len(results)}

    except Exception as e:
        return {"error": f"Error reading file: {str(e)}"}

def generate_svg_ul(params: dict) -> str:
    """UL Delay Window SVG 생성."""
    W = 1200
    LEFT_PAD = 180
    PAD_TOP = 40
    ROW_H = 40
    
    # 축 범위 계산
    all_values = []
    for key in ['T1a_max_cp_ul', 'T1a_min_cp_ul', 'Ta4_max_ul', 'Ta4_min_ul',
                'T2a_max_cp_ul', 'T2a_min_cp_ul', 'Ta3_max_ul', 'Ta3_min_ul',
                'real_T1a_max_cp_ul', 'real_T1a_min_cp_ul', 'real_Ta4_max_ul', 'real_Ta4_min_ul',
                'real_T2a_max_cp_ul', 'real_T2a_min_cp_ul', 'real_Ta3_max_ul', 'real_Ta3_min_ul']:
        val = nnum(params.get(key))
        if val is not None:
            all_values.append(val)
    
    if not all_values:
        minX, maxX = -800, 800
    else:
        minX, maxX = min(all_values), max(all_values)
        span = maxX - minX
        if span == 0:
            span = 100
        pad = span * 0.10
        minX -= pad
        maxX += pad
    
    def scale(x):
        return LEFT_PAD + (x - minX) / (maxX - minX) * (W - LEFT_PAD - 80)
    
    pairs = [
        ("O-DU", "Tx Win C-P",      "T1a_max_cp_ul", "T1a_min_cp_ul",      "real_T1a_max_cp_ul", "real_T1a_min_cp_ul", False, "T1a"),
        ("O-DU", "Rx Win U-P",      "Ta4_max_ul",    "Ta4_min_ul",         "real_Ta4_max_ul",    "real_Ta4_min_ul", True, "Ta4"),
        ("O-RU", "Rx Win C-P",      "T2a_max_cp_ul", "T2a_min_cp_ul",      "real_T2a_max_cp_ul", "real_T2a_min_cp_ul", False, "T2a"),
        ("O-RU", "Tx Win U-P",      "Ta3_max_ul",    "Ta3_min_ul",         "real_Ta3_max_ul",    "real_Ta3_min_ul", True, "Ta3")
    ]
    
    H = PAD_TOP + len(pairs) * (ROW_H * 2 + 18) + 40
    parts = []
    parts.append(f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" class="w-full border rounded-xl bg-white">')
    
    # 배경 그리드 및 축
    parts.append(f'<g transform="translate(0,{PAD_TOP})">')
    parts.append(f'<line x1="{LEFT_PAD}" x2="{W-80}" y1="0" y2="0" stroke="#cbd5e1" stroke-width="1"/>')
    
    # 그리드 라인 및 레이블
    for i in range(5):
        t = i / 4.0
        xval = minX + t * (maxX - minX)
        x = scale(xval)
        parts.append(f'<line x1="{x:.2f}" x2="{x:.2f}" y1="0" y2="{H-70}" stroke="#e2e8f0" stroke-width="1"/>')
        parts.append(f'<text x="{x:.2f}" y="-8" text-anchor="middle" font-size="11" fill="#475569">{fmt_ns(xval)}</text>')
    
    # 중앙선 (t=0)
    x0 = scale(0.0)
    parts.append(f'<line x1="{x0:.2f}" x2="{x0:.2f}" y1="0" y2="{H-70}" stroke="#94a3b8" stroke-dasharray="4 4" stroke-width="1"/>')
    parts.append(f'<text x="{x0:.2f}" y="14" font-size="11" fill="#334155">t=0</text>')
    parts.append('</g>')
    
    # 데이터 영역
    parts.append('<g>')
    TEXT_DARK = "#0f172a"
    
    for idx, pair_info in enumerate(pairs):
        equip, lane, k_es, k_ee, k_ms, k_me, is_positive, timer_name = pair_info
        
        baseY = PAD_TOP + 18 + idx * (ROW_H * 2 + 18)
        
        # 장비 레이블 (C-P일 때만 표시)
        if "C-P" in lane:
            parts.append(f'<text x="{LEFT_PAD-40}" y="{baseY-6}" text-anchor="end" font-size="15" font-weight="600" fill="{TEXT_DARK}">{equip}</text>')
        
        # 값 가져오기
        es_raw = nnum(params.get(k_es))
        ee_raw = nnum(params.get(k_ee))
        ms_raw = nnum(params.get(k_ms))
        me_raw = nnum(params.get(k_me))
        
        # 양수/음수 처리
        if is_positive:
            ms = me_raw
            me = ms_raw
            es = es_raw
            ee = ee_raw
        else:
            ms = ms_raw
            me = me_raw
            es = es_raw
            ee = ee_raw
        
        # 색상 팔레트
        EXP_FILL, EXP_EDGE, MEA_FILL, MEA_EDGE = palette_for_row(equip, lane, params)
        
        # 평면 타입
        plane_type = "CP" if "C-P" in lane else "UP"
        
        # -------- Expected row --------
        y_exp = baseY + 10
        if es is not None and ee is not None:
            exp_min_val = min(es, ee)
            exp_max_val = max(es, ee)
            
            xs = scale(exp_min_val)
            xe = scale(exp_max_val)
            rectX = xs
            rectW = xe - xs
            
            # 예상 영역
            parts.append(f'<rect x="{rectX:.2f}" y="{y_exp}" width="{rectW:.2f}" height="24" fill="{EXP_FILL}" stroke="{EXP_EDGE}" stroke-width="1.5"/>')
            
            # 시작 마커
            parts.append(f'<line x1="{xs:.2f}" x2="{xs:.2f}" y1="{y_exp-8}" y2="{y_exp+24}" stroke="{EXP_EDGE}" stroke-width="2"/>')
            parts.append(f'<text x="{xs:.2f}" y="{y_exp-12}" text-anchor="middle" font-size="11" fill="#1e3a8a">Exp {timer_name} {plane_type} min {fmt_ns(exp_min_val)}</text>')
            
            # 끝 마커
            parts.append(f'<line x1="{xe:.2f}" x2="{xe:.2f}" y1="{y_exp-8}" y2="{y_exp+24}" stroke="{EXP_EDGE}" stroke-width="2"/>')
            parts.append(f'<text x="{xe:.2f}" y="{y_exp-12}" text-anchor="middle" font-size="11" fill="#1e3a8a">Exp {timer_name} {plane_type} max {fmt_ns(exp_max_val)}</text>')
            
            # 레이블
            parts.append(f'<text x="{LEFT_PAD-70}" y="{y_exp+18}" font-size="11" fill="{TEXT_DARK}">{lane} Exp</text>')
        else:
            parts.append(f'<text x="{LEFT_PAD-30}" y="{y_exp+16}" font-size="11" fill="#ef4444">N/A</text>')
        
        # -------- Measured row --------
        y_mea = y_exp + ROW_H
        if ms is not None and me is not None:
            meas_min_val = min(ms, me)
            meas_max_val = max(ms, me)
            
            xs_m = scale(meas_min_val)
            xe_m = scale(meas_max_val)
            rectX_m = xs_m
            rectW_m = xe_m - xs_m
            
            # 측정 영역
            parts.append(f'<rect x="{rectX_m:.2f}" y="{y_mea}" width="{rectW_m:.2f}" height="24" fill="{MEA_FILL}" stroke="{MEA_EDGE}" stroke-width="1.5"/>')
            
            # 상태 평가
            if is_positive:
                s_txt, s_col = ul_status_pos(ms, es, ee, is_start=True)
                e_txt, e_col = ul_status_pos(me, es, ee, is_start=False)
            else:
                s_txt, s_col = ul_status_neg(ms, es, ee, is_start=True)
                e_txt, e_col = ul_status_neg(me, es, ee, is_start=False)
            
            # 시작 마커
            parts.append(f'<line x1="{xs_m:.2f}" x2="{xs_m:.2f}" y1="{y_mea-8}" y2="{y_mea+24}" stroke="{s_col}" stroke-width="2"/>')
            start_label_x = xs_m - 10
            start_label_y = y_mea + 10
            parts.append(f'<text x="{start_label_x:.2f}" y="{start_label_y:.2f}" text-anchor="end" font-size="11" fill="{s_col}">Start {timer_name} {plane_type} {fmt_ns(ms)} {s_txt}</text>')
            
            # 끝 마커
            parts.append(f'<line x1="{xe_m:.2f}" x2="{xe_m:.2f}" y1="{y_mea-8}" y2="{y_mea+24}" stroke="{e_col}" stroke-width="2"/>')
            end_label_x = xe_m + 10
            end_label_y = y_mea + 10
            parts.append(f'<text x="{end_label_x:.2f}" y="{end_label_y:.2f}" font-size="11" fill="{e_col}">End {timer_name} {plane_type} {fmt_ns(me)} {e_txt}</text>')
            
            # 레이블
            parts.append(f'<text x="{LEFT_PAD-70}" y="{y_mea+18}" font-size="11" fill="{TEXT_DARK}">{lane} Meas</text>')
        else:
            parts.append(f'<text x="{LEFT_PAD-30}" y="{y_mea+16}" font-size="11" fill="#ef4444">N/A</text>')
    
    parts.append('</g>')
    parts.append('</svg>')
    return "".join(parts)

# ------------------------- 색상 관련 함수들 -------------------------
def _parse_hex_color(h: str):
    """16진수 색상 파싱"""
    if not h: return (0, 0, 0)
    s = str(h).strip().lstrip("#")
    if len(s) == 3:
        s = "".join(ch * 2 for ch in s)
    if len(s) != 6: return (0, 0, 0)
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
        return (r, g, b)
    except Exception:
        return (0, 0, 0)

def _rgb_to_hex(r, g, b):
    """RGB를 16진수로 변환"""
    r = max(0, min(255, int(round(r))))
    g = max(0, min(255, int(round(g))))
    b = max(0, min(255, int(round(b))))
    return f"#{r:02x}{g:02x}{b:02x}"

def _lighten_hex(h: str, amount: float = 0.6) -> str:
    """색상 밝게 만들기"""
    r, g, b = _parse_hex_color(h)
    r = r + (255 - r) * amount
    g = g + (255 - g) * amount
    b = b + (255 - b) * amount
    return _rgb_to_hex(r, g, b)

def _darken_hex(h: str, amount: float = 0.25) -> str:
    """색상 어둡게 만들기"""
    r, g, b = _parse_hex_color(h)
    r = r * (1 - amount)
    g = g * (1 - amount)
    b = b * (1 - amount)
    return _rgb_to_hex(r, g, b)

def is_user_plane_by_text(label: str) -> bool:
    """레이블에서 User Plane 여부 확인"""
    if not label: return False
    s = str(label).lower()
    return ("u-p" in s) or ("user" in s) or ("u plane" in s) or ("u_plane" in s)

def palette_for_row(equip: str, lane_text: str, params: dict):
    """행별 색상 팔레트 생성"""
    is_up = is_user_plane_by_text(lane_text)
    equip_upper = str(equip).upper()
    
    if equip_upper.startswith("O-DU"):
        key = "color_odu_up" if is_up else "color_odu_cp"
    else:
        key = "color_oru_up" if is_up else "color_oru_cp"
    
    base = params.get(key) or PARAM_DEFAULTS_DL.get(key) or "#60a5fa"
    if not isinstance(base, str) or not base.startswith("#"):
        base = "#60a5fa"
    
    exp_fill = _lighten_hex(base, 0.6)
    exp_edge = base
    mea_fill = base
    mea_edge = _darken_hex(base, 0.25)
    
    return exp_fill, exp_edge, mea_fill, mea_edge

# ------------------------- 엑셀 로딩 함수 -------------------------
def load_params_from_excel(file_stream, mode='dl'):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    sheet = wb.active
    params = dict(PARAM_DEFAULTS_DL if mode == 'dl' else PARAM_DEFAULTS_UL)
    
    for row in sheet.iter_rows(min_row=1, max_col=2):
        kcell, vcell = row
        key = kcell.value
        if key is None:
            continue
        key = normalize_key(str(key))
        if key in params:
            val = nnum(vcell.value)
            if val is not None:
                params[key] = val
    return params

def current_params(src, mode='dl'):
    defaults = PARAM_DEFAULTS_DL if mode == 'dl' else PARAM_DEFAULTS_UL
    params = dict(defaults)
    for k in list(params.keys()):
        if k in src:
            raw = src.get(k)
            if k.startswith("color_"):
                params[k] = str(raw)
            else:
                params[k] = nnum(raw)
    return params

# ------------------------- 라우트 -------------------------
@app.route("/")
def index():
    return render_template("index_main.html", timing_url=get_timing_app_url())

@app.route("/bba/")
def bba_index():
    """BBA Master 페이지"""
    return render_template("BBA_MASTER/Index.html")


@app.route("/dl/")
def dl_index():
    params = session.get('dl_params', None)
    if params is None or not isinstance(params, dict):
        params = dict(PARAM_DEFAULTS_DL)
    
    svg = generate_svg_dl(params)
    evals, counters, log_text, rx_checks = evaluate_dl(params)
    return render_template("index.html", params=params, svg=svg, evals=evals, counters=counters, log_text=log_text, rx_checks=rx_checks, mode='dl')

@app.route("/dl/update", methods=["POST"])
def dl_update():
    params = session.get('dl_params', dict(PARAM_DEFAULTS_DL))
    if not isinstance(params, dict):
        params = dict(PARAM_DEFAULTS_DL)

    file = request.files.get("file")

    if file and file.filename:
        params = load_params_from_excel(file.stream, 'dl')
        form_params = current_params(request.form, 'dl')
        for k, v in form_params.items():
            if k.startswith("color_"):
                params[k] = v
    else:
        form_params = current_params(request.form, 'dl')
        params.update(form_params)

    session['dl_params'] = params
    session.modified = True

    return redirect(url_for("dl_index"))

@app.route("/ul/")
def ul_index():
    params = session.get('ul_params', None)
    if params is None or not isinstance(params, dict):
        params = dict(PARAM_DEFAULTS_UL)
    
    svg = generate_svg_ul(params)
    evals, counters, log_text, rx_checks = evaluate_ul(params)
    return render_template("index_ul.html", params=params, svg=svg, evals=evals, counters=counters, log_text=log_text, rx_checks=rx_checks, mode='ul')

@app.route("/ul/update", methods=["POST"])
def ul_update():
    params = session.get('ul_params', dict(PARAM_DEFAULTS_UL))
    if not isinstance(params, dict):
        params = dict(PARAM_DEFAULTS_UL)

    file = request.files.get("file")

    if file and file.filename:
        params = load_params_from_excel(file.stream, 'ul')
        form_params = current_params(request.form, 'ul')
        for k, v in form_params.items():
            if k.startswith("color_"):
                params[k] = v
    else:
        form_params = current_params(request.form, 'ul')
        params.update(form_params)

    session['ul_params'] = params
    session.modified = True

    return redirect(url_for("ul_index"))

@app.route("/ecpri/")
def ecpri_index():
    # 세션에서 이전 분석 결과 가져오기
    analysis_results = session.get('ecpri_results', None)
    basic_stats = session.get('ecpri_basic_stats', None)
    error = session.get('ecpri_error', None)
    
    return render_template("index_iq.html", 
                         results=analysis_results, 
                         basic_stats=basic_stats, 
                         error=error)

@app.route("/ecpri/analyze", methods=["POST"])
def ecpri_analyze():
    """eCPRI 데이터 분석 엔드포인트"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
            temp_path = tmp.name
        file.save(temp_path)

        analysis_results = analyze_ecpri_data(temp_path)

        basic_stats = None
        if "error" not in analysis_results:
            try:
                basic_stats, _ = build_basic_csv_stats(temp_path)
            except Exception as e:
                analysis_results['stats_error'] = f"Failed to load CSV for basic stats: {str(e)}"

        session['ecpri_results'] = analysis_results
        session['ecpri_basic_stats'] = basic_stats
        session['ecpri_error'] = analysis_results.get('error')
        session.modified = True

        return jsonify({
            'success': 'error' not in analysis_results,
            'results': analysis_results,
            'basic_stats': basic_stats,
            'error': analysis_results.get('error')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Analysis failed: {str(e)}'}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

@app.route("/ecpri/export", methods=["POST"])

def ecpri_export():
    """eCPRI 분석 결과 Excel로 내보내기"""
    try:
        results = session.get('ecpri_results', {})
        
        if not results or 'error' in results:
            return jsonify({'error': 'No valid analysis results to export'}), 400
        
        # Excel 파일 생성
        excel_file, filename = save_ecpri_results_to_excel(results)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

# ------------------------- IQ Data Analyzer 호환 라우트 -------------------------
@app.route("/iq/")
def iq_index():
    return redirect(url_for('ecpri_index'))

@app.route("/iq/analyze", methods=["POST"])
def iq_analyze():
    return ecpri_analyze()

@app.route("/iq/export", methods=["POST"])
def iq_export():
    return ecpri_export()

@app.route("/dl/diagram.png", methods=["POST"])

def dl_diagram_png():
    params = current_params(request.form, 'dl')
    svg = generate_svg_dl(params)
    buf = BytesIO(svg.encode("utf-8"))
    buf.seek(0)
    fname = f"dl_delay_diagram_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.svg"
    return send_file(buf, mimetype="image/svg+xml", as_attachment=True, download_name=fname)

@app.route("/dl/report.pdf", methods=["POST"])
def dl_report_pdf():
    params = current_params(request.form, 'dl')
    svg = generate_svg_dl(params)
    evals, counters, log_text, rx_checks = evaluate_dl(params)
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 36
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "DL Delay Window Report")
    c.setFont("Helvetica", 9)
    now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.drawRightString(width - 40, y, now_str)
    y -= 16
    drawing = svg2rlg(BytesIO(svg.encode("utf-8")))
    dw, dh = drawing.width, drawing.height
    max_w = width - 80
    max_h = 220
    scale = min(max_w/dw, max_h/dh) if dw > 0 and dh > 0 else 1.0
    drawing.scale(scale, scale)
    new_w = dw * scale
    new_h = dh * scale
    x_chart = (width - new_w) / 2.0
    y_chart = y - new_h
    renderPDF.draw(drawing, c, x_chart, y_chart)
    y = y_chart - 10
    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"PASS: {counters['pass']} FAIL: {counters['fail']} On-time: {counters['ontime']} EARLY: {counters['early']} LATE: {counters['late']}")
    y -= 18
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "O-RU RX Window Check")
    y -= 14
    col_w = (width - 80) / 4.0
    base_x = 40
    headers = ["Plane", "On-Time", "Early", "Late"]
    c.setFont("Helvetica-Bold", 8)
    x = base_x
    for htxt in headers:
        c.drawString(x + 2, y, htxt)
        x += col_w
    y -= 8
    c.line(base_x, y, base_x + col_w * 4, y)
    y -= 10
    c.setFont("Helvetica", 8)
    def draw_rx_row(plane_label, on_key, early_key, late_key, y_row):
        c.setFillColor(colors.black)
        c.drawString(base_x + 2, y_row, plane_label)
        cells = [
            rx_checks.get(on_key, ""),
            rx_checks.get(early_key, ""),
            rx_checks.get(late_key, "")
        ]
        for idx, val in enumerate(cells, start=1):
            if not val: continue
            c.setFillColor(colors.HexColor("#22c55e") if val == "PASS" else colors.HexColor("#ef4444"))
            cx = base_x + col_w * idx + col_w / 2.0
            cy = y_row + 2
            c.circle(cx, cy, 2.5, stroke=0, fill=1)
            c.setFillColor(colors.black)
    row_h = 12
    draw_rx_row("U-Plane", "RX_On_Time", "RX_Early", "RX_Late", y)
    y -= row_h
    draw_rx_row("C-Plane", "RX_On_Time_C", "RX_Early_C", "RX_Late_C", y)
    y -= row_h + 8
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "Detailed Checks")
    y -= 14
    headers_dc = ["Window", "Edge", "Measured (ns)", "Expected (ns)", "Result"]
    colw_dc = [170, 35, 80, 80, 80]
    c.setFont("Helvetica-Bold", 8)
    x = 40
    for i, txt in enumerate(headers_dc):
        c.drawString(x + 2, y, txt)
        x += colw_dc[i]
    y -= 8
    c.line(40, y, 40 + sum(colw_dc), y)
    y -= 10
    c.setFont("Helvetica", 7)
    for ev in evals:
        for edge, meas_key, exp_key, status_key in [
            ("Start", "ms", "es", "start_status"),
            ("End", "me", "ee", "end_status"),
        ]:
            if y < 40: break
            x = 40
            c.drawString(x + 2, y, ev["label"])
            x += colw_dc[0]
            c.drawString(x + 2, y, edge)
            x += colw_dc[1]
            meas_val = ev[meas_key]
            c.drawRightString(x + colw_dc[2] - 2, y, fmt_cell(meas_val))
            x += colw_dc[2]
            exp_val = ev[exp_key]
            c.drawRightString(x + colw_dc[3] - 2, y, fmt_cell(exp_val))
            x += colw_dc[3]
            c.drawString(x + 2, y, ev[status_key])
            y -= 10
    c.showPage()
    c.save()
    buffer.seek(0)
    fname = f"dl_delay_report_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=fname)

@app.route("/ul/diagram.png", methods=["POST"])
def ul_diagram_png():
    params = current_params(request.form, 'ul')
    svg = generate_svg_ul(params)
    buf = BytesIO(svg.encode("utf-8"))
    buf.seek(0)
    fname = f"ul_delay_diagram_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.svg"
    return send_file(buf, mimetype="image/svg+xml", as_attachment=True, download_name=fname)

@app.route("/ul/report.pdf", methods=["POST"])
def ul_report_pdf():
    params = current_params(request.form, 'ul')
    svg = generate_svg_ul(params)
    evals, counters, log_text, rx_checks = evaluate_ul(params)
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 36
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "UL Delay Window Report")
    c.setFont("Helvetica", 9)
    now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.drawRightString(width - 40, y, now_str)
    y -= 16
    drawing = svg2rlg(BytesIO(svg.encode("utf-8")))
    dw, dh = drawing.width, drawing.height
    max_w = width - 80
    max_h = 220
    scale = min(max_w/dw, max_h/dh) if dw > 0 and dh > 0 else 1.0
    drawing.scale(scale, scale)
    new_w = dw * scale
    new_h = dh * scale
    x_chart = (width - new_w) / 2.0
    y_chart = y - new_h
    renderPDF.draw(drawing, c, x_chart, y_chart)
    y = y_chart - 10
    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"PASS: {counters['pass']} FAIL: {counters['fail']} On-time: {counters['ontime']} EARLY: {counters['early']} LATE: {counters['late']}")
    y -= 18
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "O-RU RX Window Check")
    y -= 14
    col_w = (width - 80) / 4.0
    base_x = 40
    headers = ["Plane", "On-Time", "Early", "Late"]
    c.setFont("Helvetica-Bold", 8)
    x = base_x
    for htxt in headers:
        c.drawString(x + 2, y, htxt)
        x += col_w
    y -= 8
    c.line(base_x, y, base_x + col_w * 4, y)
    y -= 10
    c.setFont("Helvetica", 8)
    def draw_rx_row(plane_label, on_key, early_key, late_key, y_row):
        c.setFillColor(colors.black)
        c.drawString(base_x + 2, y_row, plane_label)
        cells = [
            rx_checks.get(on_key, ""),
            rx_checks.get(early_key, ""),
            rx_checks.get(late_key, "")
        ]
        for idx, val in enumerate(cells, start=1):
            if not val: continue
            c.setFillColor(colors.HexColor("#22c55e") if val == "PASS" else colors.HexColor("#ef4444"))
            cx = base_x + col_w * idx + col_w / 2.0
            cy = y_row + 2
            c.circle(cx, cy, 2.5, stroke=0, fill=1)
            c.setFillColor(colors.black)
    row_h = 12
    #draw_rx_row("U-Plane", "RX_On_Time", "RX_Early", "RX_Late", y)
    #y -= row_h
    draw_rx_row("C-Plane", "RX_On_Time_C", "RX_Early_C", "RX_Late_C", y)
    y -= row_h + 8
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "Detailed Checks")
    y -= 14
    headers_dc = ["Window", "Edge", "Measured (ns)", "Expected (ns)", "Result"]
    colw_dc = [170, 35, 80, 80, 80]
    c.setFont("Helvetica-Bold", 8)
    x = 40
    for i, txt in enumerate(headers_dc):
        c.drawString(x + 2, y, txt)
        x += colw_dc[i]
    y -= 8
    c.line(40, y, 40 + sum(colw_dc), y)
    y -= 10
    c.setFont("Helvetica", 7)
    for ev in evals:
        for edge, meas_key, exp_key, status_key in [
            ("Start", "ms", "es", "start_status"),
            ("End", "me", "ee", "end_status"),
        ]:
            if y < 40: break
            x = 40
            c.drawString(x + 2, y, ev["label"])
            x += colw_dc[0]
            c.drawString(x + 2, y, edge)
            x += colw_dc[1]
            meas_val = ev[meas_key]
            c.drawRightString(x + colw_dc[2] - 2, y, fmt_cell(meas_val))
            x += colw_dc[2]
            exp_val = ev[exp_key]
            c.drawRightString(x + colw_dc[3] - 2, y, fmt_cell(exp_val))
            x += colw_dc[3]
            c.drawString(x + 2, y, ev[status_key])
            y -= 10
    c.showPage()
    c.save()
    buffer.seek(0)
    fname = f"ul_delay_report_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=fname)

# ------------------------- Timing Wrapper 라우트 -------------------------
@app.route("/timing/")
def timing_index():
    timing_url = get_timing_app_url(request.host)
    return render_template(
        "timing_embed.html",
        timing_url=timing_url,
        timing_host=timing_url,
    )

# ------------------------- Profile Extractor 라우트 -------------------------
@app.route("/profile/")
def profile_index():
    results = session.get('profile_results', None)
    error = session.get('profile_error', None)
    return render_template("index_profile.html", results=results, error=error)


@app.route("/profile/analyze", methods=["POST"])
@app.route("/profile/analyze/", methods=["POST"])
def profile_analyze():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.txt') as tmp:
            temp_path = tmp.name
        file.save(temp_path)

        analysis_output = extract_delay_profile_data(temp_path)

        if "error" in analysis_output:
            session['profile_results'] = None
            session['profile_error'] = analysis_output["error"]
            session.modified = True
            return jsonify({'success': False, 'error': analysis_output["error"]})

        session['profile_results'] = analysis_output["data"]
        session['profile_error'] = None
        session.modified = True
        return jsonify({'success': True, 'count': analysis_output["count"]})

    except Exception as e:
        return jsonify({'error': f'Analysis failed: {str(e)}'}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


@app.route("/profile/export", methods=["POST"])
@app.route("/profile/export/", methods=["POST"])
def profile_export():
    results = session.get('profile_results', None)
    if not results:
        return jsonify({'error': 'No profile results to export'}), 400

    selected_bandwidth = (request.form.get('bandwidth') or '').strip()
    if not selected_bandwidth:
        return jsonify({'error': 'Bandwidth is required'}), 400

    selected_row = next((row for row in results if str(row.get('bandwidth', '')).strip() == selected_bandwidth), None)
    if not selected_row:
        return jsonify({'error': 'Selected bandwidth was not found'}), 400

    export_row = {
        'Bandwidth': selected_row.get('bandwidth', ''),
        'SCS': selected_row.get('scs', ''),
        'T2a Min Up': selected_row.get('t2a-min-up', 'N/A'),
        'T2a Max Up': selected_row.get('t2a-max-up', 'N/A'),
        'T2a Min CP DL': selected_row.get('t2a-min-cp-dl', 'N/A'),
        'T2a Max CP DL': selected_row.get('t2a-max-cp-dl', 'N/A'),
        'TCP Adv DL': selected_row.get('tcp-adv-dl', 'N/A'),
        'Ta3 Min': selected_row.get('ta3-min', 'N/A'),
        'Ta3 Max': selected_row.get('ta3-max', 'N/A'),
        'T2a Min CP UL': selected_row.get('t2a-min-cp-ul', 'N/A'),
        'T2a Max CP UL': selected_row.get('t2a-max-cp-ul', 'N/A'),
    }

    df = pd.DataFrame([export_row])
    output = BytesIO()
    output.write(df.to_csv(index=False).encode('utf-8-sig'))
    output.seek(0)

    safe_bandwidth = re.sub(r'[^0-9A-Za-z._-]+', '_', selected_bandwidth)
    filename = f"profile_{safe_bandwidth}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )

@app.route("/profile/clear")
def profile_clear():
    session.pop('profile_results', None)
    session.pop('profile_error', None)
    return redirect(url_for('profile_index'))
if __name__ == "__main__":
    app.run(debug=False, host='0.0.0.0', port=5000)